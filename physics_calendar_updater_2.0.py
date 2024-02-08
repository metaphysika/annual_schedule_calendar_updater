import openpyxl
import shutil
import datetime
import time
import py, os
from zipfile import BadZipFile
from emailsender_webmail import *
import getpass

# Get password via user input to use for sending email
try:
    password = getpass.getpass()
except Exception as error:
    print('ERROR', error)
emailname2 = ["christopher.lahn@sanfordhealth.org",]
# This sends a test email.  If it can't send, the script stops.  This will prevent scripts from partially processing and failing on email sending.
try:
    EmailSender().send_email(password, emailname2, 
                             "DELETE ME PLEASE",
                             "This is just checking that the correct password was entered for this script")
    pass
except Exception as error:
    print('ERROR: Could not send test email.  Wrong password?', error)

#-----------------------------------------------------------------------------
# Script variables
#-----------------------------------------------------------------------------

# Base directory and path seperator alias
dirBase = py.path.local(os.getenv('USERPROFILE'))
dirBase = dirBase.join('Sanford Health', 
                       'Imaging Physics - Equipment Testing Library')
dirBase2 = py.path.local("W:\SHARE8 Physics")
dirBase3 = py.path.local("H:")

# List of directories in which to check for reports
dirs2Check = [dirBase.join('Surveys', 'Radiography'),
              dirBase.join('Surveys', 'Fluoroscopy'),
              dirBase2.join('General Xray'),
              dirBase2.join('Fluoroscopy')]

# list of diretcories by employee type for writing to appropriate tab on Equipment Summary Sheet
qmdir = [dirBase.join('Surveys','Radiography'), 
         dirBase.join('Surveys', 'Fluoroscopy'),
         dirBase2.join('Fluoroscopy'),
         dirBase2.join('General Xray')]
phydir = [dirBase.join('Surveys', 'CT'),
          dirBase.join('Surveys', 'Mammography'),
          dirBase.join('Surveys', 'MRI'),
          dirBase.join('Surveys', 'NM'),
          dirBase.join('Surveys', 'PET')]

# Workbook names
# wbOutput = dirBase.join('Equipment Calendar',
#                         'Upcoming Equipment Survey Summary Sheet(onedrive).xlsx')
# wbDb = dirBase.join('Equipment Calendar',
#                     'Physics Equipment List Auto Report Data(onedrive).xlsx')
wbOutput = dirBase2.join('Equipment list',
                        'Upcoming Equipment Survey Summary Sheet(onedrive).xlsx')
wbDb = dirBase2.join('Equipment list',
                    'Physics Equipment List Auto Report Data(onedrive).xlsx')                        
# wbDbBck = py.path.local(os.getenv('HOMEDRIVE')).join('Scripts',
#                                                      'Physics Calendar Updater',
#                                                      'Physics Equipment List Auto Report Data(onedrive).xlsx')
wbDbBck = dirBase3.join('Scripts',
                        'Physics Calendar Updater',
                        'Physics Equipment List Auto Report Data(onedrive).xlsx')
# wbDbBck.ensure_dir()

# Named ragnes to search for in report spreadsheets
nmsFacility = ['facility', 'Facility']  # named ranges that might contain the facility name
nmsRoom = ['room_id', 'RmID']  # named ranges that might contain the room ID
nmsSysId = ['system_id']  # named ranges for the system ID
nmsDateSur = ['survey_date', 'DateSur']  # named ranges for survey date
nmsDateRev = ['review_date']  # named ranges for survey review date

# List of keywords to exclude. Directories containing these keywords will be
# excluded from the report spreadshit search. **NOTE** case insensitive.
# **WARNING** any directory containing these keywords will be excluded. Be
# careful that hte keywords aren't to short (i.e., too inclusive)
exclude = [
    "archive",
    "backup",
    "biomed",
    "blank form",
    "carestream image look tool",
    "demo",
    "\\mee\\",
    "mini c-arm",
    "other",
    "procedure times",
    "protocol",
    "qc",
    "removed",
    "audit",
    "test exclude",
    'RWS',
    '--re']
exclude = [e.lower() for e in exclude]

# File extension filter
ext = "*.xl*"  # file extension of files to consider


def _get_name_cell_value(wb, names):
    """
    Get cell contents from openpyxl workbook

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook in which to check for named ranges

    names : list
        List of strings (cell names) to try.

    Returns
    -------
    cell_val : TYPE
        DESCRIPTION.

    """

    val = None  # Init output

    # Attempt to get output, return None if not found
    for nm in names:

        # Attempt to get the named range. A KeyError will be thrown when
        # getting the title, coord if it doesn't exist
        try:
            title, coord = next(wb.defined_names[nm].destinations)
            val = wb[title][coord].value
        except (KeyError, StopIteration, AttributeError):
            pass

        if val is not None:
            break

    return val


# This clears the workbook so it can be updated with new day's data
wb = openpyxl.load_workbook(filename=wbOutput)
ws = wb['GEN RAD']
ws.delete_rows(2,100)
# for row in ws['A1:H300']:
#     for cell in row:
#         cell.value = None
ws1 = wb['GEN RAD REPORTS']
ws1.delete_rows(2,100)
# for row in ws1['A1:H300']:
#     for cell in row:
#         cell.value = None
wb.save(wbOutput)
wb.close()

# For each sub-directory of dirs2Check, get Excel spreadsheets (filtered by
# extension). Extract data from those files
for dirChk in dirs2Check:

    # Get the Excel files
    fList = []
    for f in dirChk.visit(fil=ext):
        # Checks exclude list for subdirectories to ingore
        if any([ex in f.strpath.lower() for ex in exclude]):
            continue
        try:
            # Attempt to convert the first 8 digits of the file name to a date
            dateSurvey = datetime.datetime.strptime(f.purebasename[:8],
                                                    "%Y%m%d")
            fList.append(f)
        except ValueError:
            pass  # don't do anything with files that aren't named correctly

    # Get the unique directory names
    dList = []
    for f in fList:
        if f.parts()[-2] not in dList:
            dList.append(str(f.parts()[-2]))  # py.path.local not needed

    # For each unique directory, determine the file with the most recent date
    for d in dList:

        # Get the files corresponding to the current sub-directory
        fListUnique = [f for f in fList if d in f.strpath]

        # Get the most recent date and convert back to a string
        fDate = max([datetime.datetime.strptime(f.purebasename[:8], "%Y%m%d")
                     for f in fListUnique])
        fDate = datetime.datetime.strftime(fDate, "%Y%m%d")

        # Get the file corresponding to that date
        f = [str(f) for f in fListUnique if fDate in f.purebasename]
        if len(f) > 1:
            tNewest = max([py.path.local(f).mtime() for f in f])
            f = [f for f in f if py.path.local(f).mtime() == tNewest]
            if len(f) > 1:
                h = set([py.path.local(f).computehash() for f in f])
                if len(h) > 1:
                    # raise ImportError("I give up...")
                    emailname = ["christopher.lahn@sanfordhealth.org",]
                    # raise ImportError("I give up...")
                    EmailSender().send_email(password, emailname, "Automated Message:  Physics Calendar Updater Problem", 
                        "There was a problem with the following files: " + str(f)) 
                    continue
        f = f[0]

        # Attempt to get the data
        try:

            print(f)
            # The data_only =True is important to ignore the conditional
            # formatting errors that pop up in certain forms. I was mainly
            # seeing them on the fluoro forms.
            wb = openpyxl.load_workbook(f, data_only=True)

            # This gets the facility name
            facility = _get_name_cell_value(wb, nmsFacility)

            # this gets the name of the equipment
            equipment = _get_name_cell_value(wb, nmsRoom)

            # this gets the GE Id
            try:
                system = _get_name_cell_value(wb, nmsSysId)
            except IndexError:
                print ("no system ID found ", f)
                pass

            # this gets the review date equipment reviewed
            review = _get_name_cell_value(wb, nmsDateRev)
            try:
                # print(review)
                # Set format of date with strftime()
                # review2 = review.strftime("%Y/%m/%d")
                review2 = review.strftime("%m/%d/%Y")
            except Exception:
                review2 = None

            # this gets the date equipment tested
            result = _get_name_cell_value(wb, nmsDateSur)
            try:
                #print (result)
                # Set format of date with strftime()
                # result2 = result.strftime("%Y/%m/%d")
                result2 = result.strftime("%m/%d/%Y")
            except Exception:
                result2 = None

            # This subtracts date and gives a date 1 year prior
            # print (result - datetime.timedelta(days=365))
            # this calculates how many days since date on report
            try:
                overdue = datetime.datetime.now() - result
                print(overdue)
            except (StopIteration, AttributeError):
                overdue = None
            except TypeError:
                print(f)


            # Check if report for upcoming survey.  If blank survey date, pass.
            if result2 == None:
                print ('no survey date - or upcoming survey')
                pass
            else:
                # This will open the testing calendar, find the equipment and update the datescl
                gencal = openpyxl.load_workbook(wbDb)
                if dirChk in qmdir:
                    ws = gencal["Gen Rad"]
                elif dirChk in phydir:
                    ws = gencal["Physicist"]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == system and cell.value != None:
                            # print (result)
                            compdate = ws.cell(row=cell.row, column=2).value
                            # print (compdate)
                            stringdate = datetime.datetime.strptime(compdate, "%m/%d/%Y")
                            # print (stringdate)
                            # This checks if the opened report is older than the last recorded testings date.  If so, it goes to next file.
                            if stringdate <= result:
                                # the matches the name of the equipment on the report with the equip. on the calendar
                                if cell.value == system and cell.value != None:
                                    # This writes the date from the report to the column next to the equip on the calendar.
                                    ws.cell(row=cell.row, column=2, value=result2)
                                    ws.cell(row=cell.row, column=3, value=review2)
                                    ws.cell(row=cell.row, column=4, value=facility)
                                    ws.cell(row=cell.row, column=8, value=f)
                                    # ws.cell(row=cell.row, column=5, value=system) #Don't need to write the system id because it is already there?
                                    region = (ws.cell(row=cell.row, column=9).value)
                                    gencal.save(wbDb)

                                    # This will check if the equipment on report is due or late.
                                    # Late = reports with dates > 365 days
                                    if overdue > datetime.timedelta(days=385):
                                        print('I am late!')
                                        dates = []
                                        dates.append(facility)
                                        dates.append(equipment)
                                        dates.append(result2)
                                        dates.append("late")
                                        dates.append(system)
                                        dates.append(f)
                                        dates.append(region)
                                        # print(dates)
                                        genxmaster = openpyxl.load_workbook(wbOutput)
                                        # sheet = genxmaster.active # used for defualt sheet
                                        if dirChk in qmdir:
                                            sheet = genxmaster["GEN RAD"]  # can specify which sheet it writes to.
                                            sheet.append(dates)
                                        elif dirChk in phydir:
                                            sheet = genxmaster["PHYSICIST"]  # can specify which sheet it writes to.
                                            sheet.append(dates)
                                        # for row, text in enumerate(dates, start=2):
                                        #     ws.cell(column=1, row=row, value=text)
                                        genxmaster.save(wbOutput)
                                        genxmaster.close()
                                        #print ("true", overdue)
                                    # Due = reports with dates > 335 days but < 365 days.
                                    elif overdue > datetime.timedelta(days=335) and overdue <= datetime.timedelta(days=385):
                                        dates = []
                                        dates.append(facility)
                                        dates.append(equipment)
                                        dates.append(result2)
                                        dates.append("due soon")
                                        dates.append(system)
                                        dates.append(f)
                                        dates.append(region)
                                        genxmaster = openpyxl.load_workbook(wbOutput)
                                        if dirChk in qmdir:
                                            sheet = genxmaster["GEN RAD"]  # can specify which sheet it writes to.
                                            sheet.append(dates)
                                        elif dirChk in phydir:
                                            sheet = genxmaster["PHYSICIST"]  # can specify which sheet it writes to.
                                            sheet.append(dates)
                                        genxmaster.save(wbOutput)
                                        genxmaster.close()
                                    else:
                                        #print ("false", overdue)
                                        pass

                                    # this creates list of reports that need to be reviewed
                                    if overdue < datetime.timedelta(days=31) and review == None:
                                        report = []
                                        report.append(facility)
                                        report.append(equipment)
                                        report.append(result2)
                                        report.append("due soon")
                                        report.append(system)
                                        report.append(f)
                                        report.append(region)
                                        genxmaster = openpyxl.load_workbook(wbOutput)
                                        if dirChk in qmdir:
                                            sheet = genxmaster["GEN RAD REPORTS"]
                                            sheet.append(report)
                                        if dirChk in phydir:
                                            pass
                                        genxmaster.save(wbOutput)
                                        genxmaster.close()
                                        #print("report due", overdue)
                                    elif overdue > datetime.timedelta(days=31) and review == None:
                                        report = []
                                        report.append(facility)
                                        report.append(equipment)
                                        report.append(result2)
                                        report.append("late")
                                        report.append(system)
                                        report.append(f)
                                        report.append(region)
                                        genxmaster = openpyxl.load_workbook(wbOutput)
                                        if dirChk in qmdir:
                                            sheet = genxmaster["GEN RAD REPORTS"]
                                            sheet.append(report)
                                        if dirChk in phydir:
                                            pass
                                        genxmaster.save(wbOutput)
                                        genxmaster.close()
                                        #print("report late")
                                    else:
                                        #print ("not late reoprt", overdue)
                                        pass
                                    time.sleep(1)
                            else:
                                # just a check to display if this is a report in the archive folders that is not the most recent report.
                                print ("Not new report")
                                pass

            # wb.save(r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
            # wb.close()
        # This ensures the iterator will keep going through the folders looking for
        # .xlsx files, even if it finds a folder with none in itself.
        # Without the try/except setup, the code throws an error upon finding a folder without
        # any .xlsx files and quits searching in other folders (sort of, it acutally quits storing the files in the max variable).
        except (ValueError, IOError) as error:
            print(error)
            pass
        except (TypeError) as error:
            print(f, error)
        except (BadZipFile) as error:
            print(f, error)

# copy backup of autoreport data to H drive
# shutil.copy(wbDb, wbDbBck)
# TODO: Email out a copy of the report daily/weekly to everyone?
# TODO: Maybe figure out how to add due items as tasks to the Physics calendar?
