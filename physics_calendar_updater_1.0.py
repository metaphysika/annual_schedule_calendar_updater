import openpyxl
import datetime
import py
import pandas as pd
import xlwings as xw
import xlsxwriter
import os
import string
from zipfile import BadZipFile

# TODO:  Point the crawler at Physics drive General Xray folder
# TODO: Create second script or add to this script that does this for fluoro


# genxmaster = openpyxl.load_workbook("W:\SHARE8 Physics\Equipment list\Test QM Equipment List Current Year.xlsx")
# This clears the workbook so it can be updated with new day's data
wb = openpyxl.load_workbook(
    filename=r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
ws = wb['GEN RAD']
for row in ws['A1:H300']:
    for cell in row:
        cell.value = None
ws1 = wb['GEN RAD REPORTS']
for row in ws1['A1:H300']:
    for cell in row:
        cell.value = None
wb.save(r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
wb.close()

# This sets the header row on each sheet
wb = openpyxl.load_workbook(
    filename=r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
ws = wb['GEN RAD']
ws1 = wb['GEN RAD REPORTS']
header = ["Facility", "Equipment", "Last Survey Date", "Status", "GE ID"]
ws.append(header)
ws1.append(header)
wb.save(r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
wb.close()


# TODO:  add directories you want to exclude here.
# This is the exclude list for folders you want os.walk to ignore
exclude = [
    "text exclude", "1 X-RAY BLANK FORMS", "2 ARCHIVED - OLD DOCUMENTS",
    "3 REMOVED FROM SERVICE", "4 DEMO SURVEYS", "1 FLUORO BLANK FORMS",
    "1 BISMARCK ARCHIVED - OLD DOCUMENTS", "Biomed", "FARGO ARCHIVE", "Other",
    "1 SMCF ARCHIVED - OLD DOCUMENTS", "1 ARCHIVED", "BISMARCK QC",
    "Carestream Image Look Tool", "MANDAN QC", "4 FLUORO PROCEDURE TIMES",
    "5 Protocols", "6 Backups", "Mini C-Arm"
]


def extract_number(f):
    try:
        s = (int(os.path.splitext(os.path.basename(f))[0].split('_')[0]))
        return s
    except Exception:
        # if the above criteria isn't met (no '_' in file name) then it returns a value of 1.
        # the 1 value will allow the max value to never select these files that do not meet naming criteria.
        return 1

    # return (s if s else -1, f)


# This walks through all the subdirectories looking for .xlsx files of a specific diretory pointed at.
# I pointed it at two different directories (for now will be gen rad and fluro).  Will add CT, MRI, Mammo later.
ext = (".xlsx", "xlsm")
ign = ("$", "~")
for x in (r'W:\SHARE8 Physics\General x-ray units', r'W:\SHARE8 Physics\Fluoroscopy'):
    # for x in (r'W:\SHARE8 Physics\General x-ray units\5 TestReportScheduler', r'W:\SHARE8 Physics\Fluoroscopy\7 Test Fluoro Calendar Update'):
    # for x in (r'W:\SHARE8 Physics\General x-ray units'):
    # for root, dirs, files in os.walk(r'W:\SHARE8 Physics\General x-ray units\5 TestReportScheduler'):
    for root, dirs, files in os.walk(x):
        # checks exclude list for subdirectories to ingore in os.walk
        dirs[:] = [d for d in dirs if d not in exclude]
        # list that will hold the files in each subdirectory to then be checked for file with highest date in name.

        # this was ryans list comprehension version.  I didn't know where to put the os.path.join.
        #list_of_files = [f for f in files if f.endswith(ext) and not f.startswith(ign)]
        list_of_files = []
        for file in files:
            if file.endswith(tuple(ext)) and not file.startswith(ign):
                # this appends the file name and full path to list_of_files
                list_of_files.append(os.path.join(root, file))
        '''
        try:
            file_number = int(os.path.splitext(os.path.basename(file))[0].split()[0])
            list_of_files.append(file_number)
        except ValueError:
            pass
        '''
        try:
            # extract number is a function that finds the largest number in the list_of_files and returns that file.
            largest = max(list_of_files, key=extract_number)
            # largest = max(list_of_files)
            # print(max(list_of_files))
            # The data_only =True is important to ignore the conditional formatting errors that pop up in certain forms.
            # I was mainly seeing them on the fluoro forms.
            wb = openpyxl.load_workbook(largest, data_only=True)
            # print (wb.sheetnames)
            # sheet = wb['Sheet2']

            # This gets the facility name
            try:
                title, coord = next(wb.defined_names['facility'].destinations)
                facility = wb[title][coord].value
                #print (facility)
            except KeyError:
                pass  # Tell the user that "review_date" name doesn't exist
            except (StopIteration, AttributeError):
                facility = None

            # this gets the name of the equipment
            try:
                title, coord = next(wb.defined_names['room_id'].destinations)
                equipment = wb[title][coord].value
                #print (equipment)
            except KeyError:
                pass  # Tell the user that "review_date" name doesn't exist
            except (StopIteration, AttributeError):
                equipment = None

            # this gets the GE Id
            try:
                title, coord = next(wb.defined_names['system_id'].destinations)
                system = wb[title][coord].value
                #print (system)
            except KeyError:
                pass  # Tell the user that "review_date" name doesn't exist
            except (StopIteration, AttributeError):
                system = None

            # this gets the review date equipment last tested
            try:
                title, coord = next(wb.defined_names['review_date'].destinations)
                review = wb[title][coord].value
                # print(review)
                # Set format of date with strftime()
                # review2 = review.strftime("%Y/%m/%d")
                review2 = review.strftime("%m/%d/%Y")
            except KeyError:
                pass  # Tell the user that "review_date" name doesn't exist
            except (StopIteration, AttributeError):
                review = None
                review2 = None

            # this gets the date equipment last tested
            try:
                title, coord = next(wb.defined_names['survey_date'].destinations)
                result = wb[title][coord].value
                #print (result)
                # Set format of date with strftime()
                # result2 = result.strftime("%Y/%m/%d")
                result2 = result.strftime("%m/%d/%Y")
            except KeyError:
                pass  # Tell the user that "review_date" name doesn't exist
            except (StopIteration, AttributeError):
                result = None
                result2 = None

            # This subtracts date and gives a date 1 year prior
            # print (result - datetime.timedelta(days=365))
            # this calculates how many days since date on report
            try:
                overdue = datetime.datetime.now() - result
            except (StopIteration, AttributeError):
                overdue = None
            except TypeError:
                print(largest)

            # This will open the testing calendar, find the equipment and update the dates
            gencal = openpyxl.load_workbook(
                r"W:\SHARE8 Physics\Equipment list\Physics Equipment List Auto Report Data.xlsx")
            ws = gencal["Gen Rad"]
            for row in ws.iter_rows():
                for cell in row:
                    # the matches the name of the equipment on the report with the equip. on the calendar
                    if cell.value == system and cell.value != None:
                        # This writes the date from the report to the column next to the equip on the calendar.
                        ws.cell(row=cell.row, column=2, value=result2)
                        ws.cell(row=cell.row, column=3, value=review2)
                        ws.cell(row=cell.row, column=4, value=facility)
                        # ws.cell(row=cell.row, column=5, value=system) #Don't need to write the system id because it is already there?
            gencal.save("W:\SHARE8 Physics\Equipment list\Physics Equipment List Auto Report Data.xlsx")

            # This will check if the equipment on report is due or late.
            # Late = reports with dates > 365 days
            if overdue > datetime.timedelta(days=365):
                dates = []
                dates.append(facility)
                dates.append(equipment)
                dates.append(result2)
                dates.append("late")
                dates.append(system)
                dates.append(largest)
                genxmaster = openpyxl.load_workbook(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                # sheet = genxmaster.active # used for defualt sheet
                sheet = genxmaster["GEN RAD"]  # can specify which sheet it writes to.
                sheet.append(dates)
                genxmaster.save(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                genxmaster.close()
                #print ("true", overdue)
            # Due = reports with dates > 335 days but < 365 days.
            elif overdue > datetime.timedelta(days=335) and overdue <= datetime.timedelta(days=365):
                dates = []
                dates.append(facility)
                dates.append(equipment)
                dates.append(result2)
                dates.append("due soon")
                dates.append(system)
                dates.append(largest)
                genxmaster = openpyxl.load_workbook(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                sheet = genxmaster["GEN RAD"]
                sheet.append(dates)
                genxmaster.save(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                genxmaster.close()
            else:
                #print ("false", overdue)
                pass

            # this creates list of reports that need to be reviewed
            if overdue < datetime.timedelta(days=31) and review == None:
                report = []
                report.append(facility)
                report.append(equipment)
                report.append(result2)
                report.append("due soon")
                report.append(system)
                report.append(largest)
                genxmaster = openpyxl.load_workbook(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                sheet = genxmaster["GEN RAD REPORTS"]
                sheet.append(report)
                genxmaster.save(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                genxmaster.close()
                #print("report due", overdue)
            elif overdue > datetime.timedelta(days=31) and review == None:
                report = []
                report.append(facility)
                report.append(equipment)
                report.append(result2)
                report.append("late")
                report.append(system)
                report.append(largest)
                genxmaster = openpyxl.load_workbook(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                sheet = genxmaster["GEN RAD REPORTS"]
                sheet.append(report)
                genxmaster.save(
                    r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
                genxmaster.close()
                #print("report late")
            else:
                #print ("not late reoprt", overdue)
                pass

            # wb.save(r'W:\SHARE8 Physics\Equipment list\Upcoming Equipment Survey Summary Sheet.xlsx')
            # wb.close()
        # This ensures the iterator will keep going through the folders looking for
        # .xlsx files, even if it finds a folder with none in itself.
        # Without the try/except setup, the code throws an error upon finding a folder without
        # any .xlsx files and quits searching in other folders (sort of, it acutally quits storing the files in the max variable).
        except (ValueError, IOError) as error:
            print(error)
            pass
        except (TypeError) as error:
            print(largest, error)
        except (BadZipFile) as error:
            print(largest, error)

# TODO: Email out a copy of the report daily/weekly to everyone?
# TODO: Maybe figure out how to add due items as tasks to the Physics calendar?
